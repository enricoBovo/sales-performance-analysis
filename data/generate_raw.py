import random
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

random.seed(42)

# ── REFERENCE DATA ────────────────────────────────────────────────────────────

CUSTOMERS = [
    "Acme Corp", "Global Tech", "Sunshine Bakery", "Rivera & Sons",
    "Blue Ridge LLC", "Novex Industries", "Peak Solutions", "Sunset Catering",
    "Bright Horizons", "Delta Supplies", "Metro Office Co", "Greenleaf Ltd"
]

EMAILS = {
    "Acme Corp":        "billing@acme.com",
    "Global Tech":      "info@globaltech.com",
    "Sunshine Bakery":  "orders@sunshine.com",
    "Rivera & Sons":    "contact@riverasons.com",
    "Blue Ridge LLC":   "admin@blueridge.com",
    "Novex Industries": "purchase@novex.com",
    "Peak Solutions":   "info@peaksol.com",
    "Sunset Catering":  "orders@sunsetcatering.com",
    "Bright Horizons":  "hello@brighthorizons.com",
    "Delta Supplies":   "orders@deltasupplies.com",
    "Metro Office Co":  "buy@metrooffice.com",
    "Greenleaf Ltd":    "contact@greenleaf.com",
}

PRODUCTS = {
    "Electronics":     [("Wireless Mouse", 29.99), ("USB-C Hub", 49.99), ("Webcam HD", 35.00),
                        ("Mechanical Keyboard", 79.99), ('Monitor 27"', 220.00), ("Laptop Stand", 45.00)],
    "Office Supplies": [("Desk Organizer", 12.50), ("Notebook Set", 18.00), ("Pen Pack", 8.99),
                        ("Desk Lamp", 34.99), ("Stapler Pro", 22.50)],
    "Furniture":       [("Ergonomic Chair", 325.00), ("Standing Desk", 549.00), ("Monitor Arm", 89.99)],
    "Software":        [("Antivirus 1yr", 59.99), ("PDF Editor", 79.00), ("VPN Annual", 49.00)],
}

REGIONS   = ["North", "South", "East", "West"]
SALES_REP = ["Sarah J.", "Mike R.", "Lisa C.", "Tom B.", "Anna K."]

# ── MESS FUNCTIONS ────────────────────────────────────────────────────────────

DATE_FORMATS = [
    lambda d: d.strftime("%Y-%m-%d"),          # 2024-01-05
    lambda d: d.strftime("%d/%m/%Y"),           # 05/01/2024
    lambda d: d.strftime("%m/%d/%Y"),           # 01/05/2024
    lambda d: d.strftime("%B %d, %Y"),          # January 05, 2024
    lambda d: d.strftime("%b %d %Y"),           # Jan 05 2024
    lambda d: d.strftime("%d-%m-%Y"),           # 05-01-2024
    lambda d: d.strftime("%Y/%m/%d"),           # 2024/01/05
    lambda d: d.strftime("%B %d %Y"),           # January 05 2024
]

def messy_date(d):
    return random.choice(DATE_FORMATS)(d)

def messy_customer(name):
    r = random.random()
    if r < 0.15: return name.upper()
    if r < 0.30: return name.lower()
    if r < 0.40: return f"  {name}  "
    if r < 0.50: return name.upper().replace(" ", "  ")
    return name

def messy_category(cat):
    typos = {"Electronics": ["Electronisc", "electronisc", "electronics", "ELECTRONICS"],
             "Office Supplies": ["office supplies", "Office supplies", "OFFICE SUPPLIES"],
             "Furniture": ["Furnitures", "furniture", "FURNITURE"],
             "Software": ["software", "SOFTWARE", "Sofware"]}
    if random.random() < 0.25 and cat in typos:
        return random.choice(typos[cat])
    return cat

def messy_price(price):
    symbols = [f"${price:.2f}", f"{price:.2f}", f"USD {price:.2f}",
               f"{price:.2f} USD", f"${price:,.2f}", str(price)]
    # Occasionally wrong currency
    if random.random() < 0.05:
        symbols += [f"£{price:.2f}", f"€{price:.2f}"]
    return random.choice(symbols)

def messy_total(total, qty, price):
    r = random.random()
    if r < 0.05: return "N/A"
    if r < 0.08: return ""
    symbols = [f"${total:.2f}", f"{total:.2f}", str(total), f"USD {total:.2f}"]
    return random.choice(symbols)

def messy_region(region):
    r = random.random()
    if r < 0.2: return region.upper()
    if r < 0.4: return region.lower()
    return region

def messy_rep(rep):
    r = random.random()
    if r < 0.2: return rep.lower()
    if r < 0.35: return rep.upper()
    return rep

# ── GENERATE ROWS ─────────────────────────────────────────────────────────────

start_date = date(2024, 1, 1)
end_date   = date(2024, 12, 31)
delta_days = (end_date - start_date).days

rows = []
for i in range(1, 1001):
    order_date  = start_date + timedelta(days=random.randint(0, delta_days))
    customer    = random.choice(CUSTOMERS)
    category    = random.choice(list(PRODUCTS.keys()))
    product, base_price = random.choice(PRODUCTS[category])
    qty         = random.randint(1, 15)
    total       = round(base_price * qty, 2)
    region      = random.choice(REGIONS)
    rep         = random.choice(SALES_REP)

    rows.append({
        "order_id":  f"ORD-{1000 + i}",
        "date":      messy_date(order_date),
        "customer":  messy_customer(customer),
        "email":     EMAILS[customer],
        "product":   product if random.random() > 0.15 else product.lower(),
        "category":  messy_category(category),
        "qty":       qty,
        "price":     messy_price(base_price),
        "total":     messy_total(total, qty, base_price),
        "region":    messy_region(region),
        "rep":       messy_rep(rep),
        "notes":     random.choice(["", "", "", "", "Urgent", "Bulk discount", "Check invoice", "Repeat customer"]),
    })

# Inject ~15 blank rows at random positions
blank = {"order_id": None, "date": None, "customer": None, "email": None,
         "product": None, "category": None, "qty": None, "price": None,
         "total": None, "region": None, "rep": None, "notes": None}
for _ in range(15):
    rows.insert(random.randint(0, len(rows)), blank.copy())

# Inject ~10 duplicate rows
for _ in range(10):
    orig = random.choice([r for r in rows if r["order_id"]])
    dup  = orig.copy()
    dup["customer"] = messy_customer(dup["customer"].strip())
    rows.insert(random.randint(0, len(rows)), dup)

# ── BUILD WORKBOOK ────────────────────────────────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = "raw_sales_data"

HEADER_BG = "4472C4"
headers = ["Order ID", "order date", "CUSTOMER NAME", "Customer_Email",
           "product", "Category", "Quantity", "Unit Price", "Total Sale",
           "Region", "Sales Rep", "  Notes  "]

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill      = PatternFill("solid", start_color=HEADER_BG, end_color=HEADER_BG)
    c.font      = Font(bold=True, color="FFFFFF", name="Arial")
    c.alignment = Alignment(horizontal="center")

for row_i, row in enumerate(rows, 2):
    vals = [row["order_id"], row["date"], row["customer"], row["email"],
            row["product"], row["category"], row["qty"], row["price"],
            row["total"], row["region"], row["rep"], row["notes"]]
    for col_i, val in enumerate(vals, 1):
        c = ws.cell(row=row_i, column=col_i, value=val)
        c.font      = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="left")

col_widths = [12, 20, 24, 30, 22, 18, 10, 16, 14, 10, 14, 22]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ── ISSUES LOG SHEET ──────────────────────────────────────────────────────────
ws2 = wb.create_sheet("data_issues_log")
ws2["A1"] = "Known Data Issues — raw_sales_data (1000 rows)"
ws2["A1"].font = Font(bold=True, size=12, name="Arial")

issues_headers = ["Issue #", "Column", "Description", "Example"]
for col, h in enumerate(issues_headers, 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.fill = PatternFill("solid", start_color=HEADER_BG, end_color=HEADER_BG)
    c.font = Font(bold=True, color="FFFFFF", name="Arial")

issues = [
    (1, "order date",    "8 different date formats mixed throughout",                  "2024-01-05 / January 05, 2024 / 05/01/2024 / Jan 05 2024 ..."),
    (2, "CUSTOMER NAME", "Inconsistent casing + extra spaces (~40% of rows affected)", "'  GLOBAL TECH  ' / 'globaltech' / 'GLOBAL TECH'"),
    (3, "product",       "~15% of product names in lowercase",                         "'wireless mouse' instead of 'Wireless Mouse'"),
    (4, "Category",      "Typos in ~25% of rows across all categories",               "'Electronisc', 'Furnitures', 'Sofware', 'office supplies'"),
    (5, "Unit Price",    "Mixed currency symbols ($, £, €) and text formats",         "'$29.99' / '29.99 USD' / 'USD 29.99' / '£29.99'"),
    (6, "Total Sale",    "~5% 'N/A', ~3% blank, rest mixed string/number format",     "'$299.90' / 299.9 / 'N/A' / ''"),
    (7, "Region",        "Inconsistent casing (~40% affected)",                        "'NORTH' / 'north' / 'North'"),
    (8, "Sales Rep",     "Inconsistent casing (~35% affected)",                        "'sarah j.' / 'SARAH J.' / 'Sarah J.'"),
    (9, "(multiple)",    "~15 blank rows inserted throughout the dataset",             "Rows with all NULL values"),
    (10,"Order ID",      "~10 duplicate order entries",                               "Same ORD-XXXX appearing twice with slight variations"),
]

for i, row_data in enumerate(issues, 4):
    for j, val in enumerate(row_data, 1):
        c = ws2.cell(row=i, column=j, value=val)
        c.font = Font(name="Arial", size=10)
        bg = "EEF2FF" if i % 2 == 0 else "FFFFFF"
        c.fill = PatternFill("solid", start_color=bg, end_color=bg)

ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 48
ws2.column_dimensions["D"].width = 52

out = "/home/claude/excel_project/raw_sales_data_1000.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total rows (incl. blanks + duplicates): {len(rows)}")
print(f"  — real orders:  1000")
print(f"  — blank rows:   ~15")
print(f"  — duplicates:   ~10")
